#!/usr/bin/env python3
"""
NBA Player Props Results Checker — BallDontLie API (Optimized Edition)
Fetches final game stats and updates your Excel sheet with ✓ / ✗ results,
including an automatic summary at the bottom.
"""

import pandas as pd
from datetime import datetime, timedelta
import time
import os
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed


# ===============================================================
# 🧠 BallDontLie Stats Fetcher Class
# ===============================================================
class NBAStatsFetcherBallDontLie:
    def __init__(self, file_path, api_key):
        self.file_path = file_path
        self.file_type = 'csv' if file_path.endswith('.csv') else 'excel'
        self.api_key = api_key
        self.base_url = "https://api.balldontlie.io/v1"
        self.team_aliases = {'NO': 'NOP', 'NY': 'NYK', 'SA': 'SAS'}
        self.player_cache = {}  # Cache player IDs to avoid re-searching
        self._test_api_connection()

    # -----------------------------------------------------------
    def _get_headers(self):
        """Use correct header format for BallDontLie API"""
        return {"Authorization": self.api_key}

    def _test_api_connection(self):
        """Ping API once to verify key"""
        headers = self._get_headers()
        try:
            r = requests.get(f"{self.base_url}/players?per_page=1", headers=headers, timeout=8)
            if r.status_code == 200:
                print("✅ BallDontLie API connected successfully!")
            elif r.status_code == 401:
                print("❌ Invalid API key — get yours at https://www.balldontlie.io/")
            else:
                print(f"⚠️ API Warning: Status {r.status_code}")
        except Exception as e:
            print(f"❌ Connection failed: {e}")

    # -----------------------------------------------------------
    def find_player_id(self, player_name):
        """Find player ID by name with caching for speed"""
        # Check cache first
        if player_name in self.player_cache:
            return self.player_cache[player_name]
        
        headers = self._get_headers()

        # Normalize name for better matching
        search_name = (
            player_name.replace("-", " ")
            .replace(".", "")
            .replace("Jr", "")
            .replace("III", "")
            .replace("II", "")
            .strip()
        )

        # Try multiple search strategies
        search_terms = [
            search_name,  # Full normalized name
            search_name.split()[0],  # First name only
            search_name.split()[-1],  # Last name only
        ]

        for search_term in search_terms:
            params = {"search": search_term, "per_page": 25}
            try:
                res = requests.get(f"{self.base_url}/players", headers=headers, params=params, timeout=10)
                
                if res.status_code != 200:
                    continue

                response_data = res.json()
                players = response_data.get("data", [])
                
                if not players:
                    continue

                # Try exact match first
                for p in players:
                    full_name = f"{p.get('first_name', '')} {p.get('last_name', '')}".strip().lower()
                    
                    if full_name == search_name.lower():
                        result = (p["id"], full_name.title())
                        self.player_cache[player_name] = result
                        return result

                # Check for partial match
                for p in players:
                    full_name = f"{p.get('first_name', '')} {p.get('last_name', '')}".strip().lower()
                    
                    if (search_name.lower() in full_name or 
                        full_name.split()[0] in search_name.lower() or 
                        full_name.split()[-1] in search_name.lower()):
                        result = (p["id"], full_name.title())
                        self.player_cache[player_name] = result
                        return result

            except Exception as e:
                continue

        # Cache the failure too to avoid re-searching
        self.player_cache[player_name] = (None, None)
        return None, None

    # -----------------------------------------------------------
    def normalize_opponent(self, opp):
        if not opp:
            return None
        return self.team_aliases.get(opp.strip().upper(), opp.strip().upper())

    # -----------------------------------------------------------
    def fetch_player_game_stats(self, player_name, opponent=None, target_date=None):
        headers = self._get_headers()
        player_id, full_name = self.find_player_id(player_name)
        if not player_id:
            return None

        if not target_date:
            target_date = datetime.now()

        # Expand search window to ensure we catch the game
        start_date = (target_date - timedelta(days=7)).strftime("%Y-%m-%d")
        end_date = (target_date + timedelta(days=2)).strftime("%Y-%m-%d")

        try:
            params = {"player_ids[]": player_id, "start_date": start_date, "end_date": end_date, "per_page": 100}
            res = requests.get(f"{self.base_url}/stats", headers=headers, params=params, timeout=10)
            if res.status_code != 200:
                return None

            games = res.json().get("data", [])
            if not games:
                return None

            # Batch fetch game details for efficiency
            game_ids = list(set([g["game"]["id"] for g in games]))
            game_details = {}
            
            for gid in game_ids:
                gr = requests.get(f"{self.base_url}/games/{gid}", headers=headers, timeout=6)
                if gr.status_code == 200:
                    game_details[gid] = gr.json()["data"]
                time.sleep(0.05)

            # Opponent-specific check
            if opponent:
                opponent = self.normalize_opponent(opponent)
                for g in games:
                    gid = g["game"]["id"]
                    game = game_details.get(gid, {})
                    if not game:
                        continue
                    
                    home = game.get("home_team", {}).get("abbreviation")
                    away = game.get("visitor_team", {}).get("abbreviation")
                    player_team = g.get("team", {}).get("abbreviation")

                    if opponent in (home, away):
                        matchup = f"{player_team} vs {away}" if player_team == home else f"{player_team} @ {home}"
                        date = g["game"]["date"].split("T")[0]
                        return {
                            "PTS": g.get("pts", 0),
                            "REB": g.get("reb", 0),
                            "AST": g.get("ast", 0),
                            "FG3M": g.get("fg3m", 0),
                            "game_date": date,
                            "matchup": matchup
                        }
                return None

            # Most recent fallback
            latest = sorted(games, key=lambda x: x["game"]["date"], reverse=True)[0]
            date = latest["game"]["date"].split("T")[0]
            home = latest["game"]["home_team"]["abbreviation"]
            away = latest["game"]["visitor_team"]["abbreviation"]
            team = latest["team"]["abbreviation"]
            matchup = f"{team} vs {away}" if team == home else f"{team} @ {home}"
            return {
                "PTS": latest.get("pts", 0),
                "REB": latest.get("reb", 0),
                "AST": latest.get("ast", 0),
                "FG3M": latest.get("fg3m", 0),
                "game_date": date,
                "matchup": matchup
            }

        except Exception as e:
            return None


    # -----------------------------------------------------------
    def process_single_prop(self, row_data, target_date):
        """Process a single prop bet (used for parallel processing)"""
        i, player, stat, line, opponent = row_data
        
        result_data = {
            'index': i,
            'player': player,
            'stat': stat,
            'line': line,
            'opponent': opponent,
            'result': None,
            'actual': None,
            'game_date': None,
            'matchup': None
        }
        
        result = self.fetch_player_game_stats(player, opponent, target_date)
        if result:
            actual = result.get(stat, 0)
            symbol = "✓" if actual > line else "✗"
            result_data['result'] = symbol
            result_data['actual'] = actual
            result_data['game_date'] = result['game_date']
            result_data['matchup'] = result['matchup']
        else:
            result_data['result'] = "⏳"
            
        return result_data

    # -----------------------------------------------------------
    def update_excel_with_results(self, target_date=None, parallel=True):
        if self.file_type == "csv":
            df = pd.read_csv(self.file_path)
        else:
            df = pd.read_excel(self.file_path)

        if target_date is None:
            target_date = datetime.now()

        print(f"\n📅 Checking results for games around {target_date.strftime('%Y-%m-%d')}")
        print(f"⚡ Mode: {'Parallel (Fast)' if parallel else 'Sequential'}\n")

        if "Result" not in df.columns:
            df["Result"] = ""

        # Prepare row data
        row_data_list = []
        for i, row in df.iterrows():
            player = row.get("Player")
            stat = row.get("Stat")
            line = float(row.get("Line", 0))
            opponent = row.get("Opponent", None)
            row_data_list.append((i, player, stat, line, opponent))

        if parallel:
            # Process in parallel for speed
            results = []
            with ThreadPoolExecutor(max_workers=5) as executor:
                future_to_row = {
                    executor.submit(self.process_single_prop, row_data, target_date): row_data 
                    for row_data in row_data_list
                }
                
                for future in as_completed(future_to_row):
                    result_data = future.result()
                    results.append(result_data)
                    
                    # Print result
                    i = result_data['index']
                    player = result_data['player']
                    stat = result_data['stat']
                    line = result_data['line']
                    opponent = result_data['opponent']
                    
                    print(f"{i+1}. {player} — {stat} {line}", end="")
                    if opponent:
                        print(f" vs {opponent}", end="")
                    
                    if result_data['result'] == "⏳":
                        print(" ⏳ Pending")
                    else:
                        print(f" → {result_data['actual']} {result_data['result']}")
            
            # Update dataframe with results
            for result_data in results:
                i = result_data['index']
                if result_data['actual'] is not None:
                    df.at[i, "Actual"] = result_data['actual']
                df.at[i, "Result"] = result_data['result']
                if result_data['game_date']:
                    df.at[i, "Game_Date"] = result_data['game_date']
                if result_data['matchup']:
                    df.at[i, "Matchup"] = result_data['matchup']
        else:
            # Sequential processing (original method)
            for i, player, stat, line, opponent in row_data_list:
                print(f"\n{i+1}. {player} — {stat} {line}")
                if opponent:
                    print(f"   🎯 vs {opponent}")

                result = self.fetch_player_game_stats(player, opponent, target_date)
                if result:
                    actual = result.get(stat, 0)
                    symbol = "✓" if actual > line else "✗"
                    df.at[i, "Actual"] = actual
                    df.at[i, "Result"] = symbol
                    df.at[i, "Game_Date"] = result["game_date"]
                    df.at[i, "Matchup"] = result["matchup"]
                    print(f"   📊 {stat}: {actual} vs {line} {symbol}")
                else:
                    df.at[i, "Result"] = "⏳"
                    print("   ⚠️ Pending or no data.")

                time.sleep(0.1)

        # Save updated file
        base, ext = os.path.splitext(self.file_path)
        output_file = f"{base}_updated{ext}"
        if ext == ".csv":
            df.to_csv(output_file, index=False)
        else:
            df.to_excel(output_file, index=False)

        print(f"\n✅ Results updated and saved → {output_file}")

        # Add summary automatically
        self.add_summary_to_excel(output_file)
        return df

    # -----------------------------------------------------------
    def add_summary_to_excel(self, output_file):
        """Append a summary section (Hits, Misses, Win %)"""
        wb = load_workbook(output_file)
        ws = wb.active

        max_row = ws.max_row + 2
        results = [cell.value for cell in ws["Result"] if cell.value in ["✓", "✗", "⏳"]]

        hits = results.count("✓")
        misses = results.count("✗")
        pending = results.count("⏳")
        total = hits + misses
        win_pct = (hits / total * 100) if total > 0 else 0

        ws.cell(row=max_row, column=1, value="Summary")
        ws.cell(row=max_row + 1, column=1, value="Hits (✓)")
        ws.cell(row=max_row + 1, column=2, value=hits)
        ws.cell(row=max_row + 2, column=1, value="Misses (✗)")
        ws.cell(row=max_row + 2, column=2, value=misses)
        ws.cell(row=max_row + 3, column=1, value="Pending (⏳)")
        ws.cell(row=max_row + 3, column=2, value=pending)
        ws.cell(row=max_row + 4, column=1, value="Total Plays")
        ws.cell(row=max_row + 4, column=2, value=total)
        ws.cell(row=max_row + 5, column=1, value="Win %")
        ws.cell(row=max_row + 5, column=2, value=f"{win_pct:.1f}%")

        wb.save(output_file)
        print(f"📊 Summary added → Hits: {hits}, Misses: {misses}, Win%: {win_pct:.1f}%\n")


# ===============================================================
# 🧩 MAIN
# ===============================================================
def main():
    API_KEY = "642d8995-44d0-4c58-b051-d32e72cf6036"
    file_path = "proppulse_results_20251112_163058.xlsx"
    target_date = datetime(2025, 11, 11)

    print("=" * 70)
    print("🏀 PropPulse+ | NBA Results Checker — BallDontLie API")
    print("=" * 70)

    fetcher = NBAStatsFetcherBallDontLie(file_path, API_KEY)
    # Set parallel=True for fast processing, parallel=False for sequential
    fetcher.update_excel_with_results(target_date=target_date, parallel=True)
# ============================================================
# ✅ Compatibility Wrapper for PropPulse+ v2025.4
# ============================================================
def fetch_player_logs(player_name, save_dir="data", refresh=False):
    """
    Wrapper for backward compatibility with prop_ev.py
    Uses NBAStatsFetcherBallDontLie to fetch player game logs.
    """
    try:
        from nba_stats_fetcher import NBAStatsFetcherBallDontLie
        fetcher = NBAStatsFetcherBallDontLie(file_path=save_dir, api_key=None)
        return fetcher.fetch_player_game_stats(player_name)
    except Exception as e:
        print(f"[Wrapper] ❌ Failed to fetch logs for {player_name}: {e}")
        return None


if __name__ == "__main__":
    main()